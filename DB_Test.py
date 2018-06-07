import cx_Oracle
from openpyxl import Workbook
import os

os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'  # 设置中文


def link_Oracle(wb):
    create_wb(wb, 'helloDatas.xlsx')
    ws = wb.create_sheet(title='我是第一个title', index=0)
    conn = cx_Oracle.connect('gpsadmin/gpsadmin_123654@ 192.168.129.135: 1521 / ora11g')  # 连接数据库
    cursor = conn.cursor()
    cursor.execute("select t.v_user_account,t.v_user_name from GPS_USER t")
    rows = cursor.fetchall()  # 得到所有数据集
    for field_desc in cursor.description:
        print(field_desc[0], end=' ,')
        # print((field_desc[0]))
    for row in rows:
        print()
        print("%s, %s" % (row[0], row[1]))
    print("Number of rows returned: %d" % cursor.rowcount)
    list_A_B = []
    rowcount1 = 1
    for a in cursor:
        rowcount1 = rowcount1 + 1
        list_A_B.append(a)
    # 读取表字段值
    db_title = [i[0] for i in cursor.description]
    # 遍历表字段值
    for titleNum in range(len(db_title)):
        A = ws.cell(row=1, column=1)
        B = ws.cell(row=1, column=2)
        A.value = db_title[titleNum]
        B.value = db_title[titleNum]
    # 读取数据到excel
    for rowNum in range(1, rowcount1):
        A = ws.cell(row=rowNum, column=1)
        B = ws.cell(row=rowNum, column=2)
        # list_A_B索引从0开始，所以-1，list中存的是tuple，后面[0]获取的是tuple中的对应元素
        A.value = list_A_B[rowNum - 1][0]
        B.value = list_A_B[rowNum - 1][1]
    cursor.close()
    conn.close()
    wb.save('helloDatas.xlsx')


# 创建Excel
def create_wb(wb, filename):
    wb.save(filename=filename)
    print("新建Excel：" + filename + "成功")


if __name__ == '__main__':
    wb = Workbook()
    link_Oracle(wb)
