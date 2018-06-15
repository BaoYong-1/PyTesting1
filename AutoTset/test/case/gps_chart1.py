# coding:utf-8
from selenium import webdriver
from xlutils3.copy import copy  # 将xlrd.Book转为xlwt.workbook，在原有的excel基础上进行修改，添加等。
import xlwt  # 写入excel（新建）
import xlrd  # 读取excel
import os
import time
import cx_Oracle
from Login import *
from openpyxl import Workbook


def load_Table(page):
    # 创建工作簿
    wbk = xlwt.Workbook(encoding='utf-8', style_compression=0)
    # 创建工作表
    sheet = wbk.add_sheet('sheet 1', cell_overwrite_ok=True)
    excel = r"F:\PyTesting\AutoTset\log\chart1.xls"
    table_rows = driver.find_element_by_xpath("//*[@id='_T202758']/td/div/div[1]/table").find_elements_by_tag_name('tr')
    row = 20
    for i, tr in enumerate(table_rows):  # enumerate()是python的内置函数.enumerate多用于在for循环中得到计数
        if i == 0 and page == 0:
            table_cols1 = tr.find_elements_by_tag_name('th')
            for j, tc in enumerate(table_cols1):
                sheet.write(i, j, tc.text)
                wbk.save(excel)
        else:
            table_cols2 = tr.find_elements_by_tag_name('td')
            for j, tc in enumerate(table_cols2):
                # 老的工作簿，打开excel
                oldWb = xlrd.open_workbook(excel, formatting_info=True)
                # 新的工作簿,复制老的工作簿
                newWb = copy(oldWb)
                # 新的工作表
                newWs = newWb.get_sheet(0)
                newWs.write(i + page * row, j, tc.text)
                os.remove(excel)
                newWb.save(excel)
    print('save done')


def switch_page():
    pages = driver.find_element_by_xpath(
        "//*[@id='_T202758']/td/div/div[3]/div/div/table/tbody/tr").find_elements_by_tag_name("td")
    t = len(pages)
    print(t)
    if t - 7 <= 9:
        for i in range(t - 7):
            driver.find_element_by_xpath(
                "//*[@id='_T202758']/td/div/div[3]/div/div/table/tbody/tr/td[3]/div/table/tbody/tr/td[" + str(
                    i + 1) + "]/div").click()
            time.sleep(3)
            driver.find_element_by_xpath(
                "//*[@id='_T202758']/td/div/div[3]/div/div/table/tbody/tr/td[3]/div/table/tbody/tr/td[" + str(
                    i + 1) + "]/div")
            load_Table(i)
    else:
        for i in range(t - 7):
            driver.find_element_by_xpath(
                "//*[@id='_T202758']/td/div/div[3]/div/div/table/tbody/tr/td[4]/div/table/tbody/tr/td[" + str(
                    i + 1) + "]/div").click()
            time.sleep(3)
            driver.find_element_by_xpath(
                "//*[@id='_T202758']/td/div/div[3]/div/div/table/tbody/tr/td[4]/div/table/tbody/tr/td[" + str(
                    i + 1) + "]/div")
            load_Table(i)


def create_wb(wb, filename):
    wb.save(filename=filename)
    print("新建Excel：" + filename + "成功")


def GetDB(user, wb):
    create_wb(wb, 'GPS_TARG.xlsx')
    ws = wb.create_sheet(title='sheet 1', index=0)
    conn = cx_Oracle.connect('gpsadmin/gpsadmin_123654@ 192.168.10.110: 1521 / ora11g')  # 连接数据库
    cursor = conn.cursor()
    cursor.execute(
        "select p.v_user_account,q.v_targ_name from GPS_USER p,GPS_TARG q where p.v_dept_id=q.v_dept_id and p.v_user_account='%s'" % user)
    rows = cursor.fetchall()  # 得到所有数据集
    for field_desc in cursor.description:
        print(field_desc[0], end=' ,')
        # print((field_desc[0]))
    for row in rows:
        print()
        print("%s, %s" % (row[0], row[1]))
    print("Number of rows returned: %d" % cursor.rowcount)
    list_A_B = []  # list
    rowcount1 = 1
    for a in rows:
        list_A_B.append(a)  # 末尾添加数据
        rowcount1 = rowcount1 + 1
    # 读取表字段值
    db_title = [i[0] for i in cursor.description]
    # 遍历表字段值
    print(len(db_title))  # 2
    for titleNum in range(len(db_title)):
        A = ws.cell(row=1, column=1)
        B = ws.cell(row=1, column=2)
        A.value = db_title[titleNum]
        B.value = db_title[titleNum]
    # 读取数据到excel
    print(rowcount1)
    for rowNum in range(1, rowcount1):
        A = ws.cell(row=rowNum, column=1)
        B = ws.cell(row=rowNum, column=2)
        # list_A_B索引从0开始，所以-1，list中存的是tuple，后面[0]获取的是tuple中的对应元素
        A.value = list_A_B[rowNum - 1][0]
        B.value = list_A_B[rowNum - 1][1]
    cursor.close()
    conn.close()
    scrpath = 'F:\\PyTesting\\AutoTset\\log\\'  # 指定的保存目录
    wb.save(scrpath + 'GPS_TARG.xlsx')


if __name__ == '__main__':
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["ignore-certificate-errors"])
    driver = webdriver.Chrome(chrome_options=options)
    driver.maximize_window()
    url = "http://192.168.10.110:8080/WebGis/login"
    driver.get(url)
    CodeText = get_code(driver)
    login(driver, 'baoyong', 'asdf1234', CodeText)
    getData(driver)
    wb = Workbook()
    switch_page()
    time.sleep(2)
    GetDB('baoyong', wb)
    login_out(driver)
    driver.quit()
