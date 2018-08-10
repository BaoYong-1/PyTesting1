# encoding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font


class Excel(object):

    def __init__(self):
        self.font = Font(color=None)
        self.colorDict = {'red': 'FFFF3030', 'green': 'FF008B00'}
        self.wb = load_workbook(test_path)
        self.ws = self.wb.active

    def rename(self, new_name):
        # 给表格重命名
        self.ws.title = new_name
        self.wb.save(test_path)

    def GetSheetName(self):
        # 获取所有表格名称
        return self.wb.get_sheet_names()

    def GetSheetByName(self, sheet_name):
        # 通过表格名称获取表格
        self.ws = self.wb.get_sheet_by_name(sheet_name)
        return self.ws

    def GetCurrentSheetName(self):
        # 获取当前表格名称
        return self.ws.title

    def GetCellContent(self, row_num, col_num):
        # 获取单元格内容
        return self.ws.cell(row=row_num, column=col_num).value

    def WriteCellContent(self, row_num, col_num, content):
        # 往指定的单元格里面写入内容
        self.ws.cell(row=row_num, column=col_num).value = content
        self.wb.save(test_path)

    def GetMaxRow(self):
        # 获取最大行号
        return self.ws.max_row

    def GetMaxColumn(self):
        # 获取最大列号
        return self.ws.max_column


if __name__ == "__main__":
    test_path = "F:\\PyTesting\\AutoTest\\log\\excel\\GPS_TARG_DB.xlsx"
    excel = Excel()
    n = excel.GetMaxRow()
    print(n)
